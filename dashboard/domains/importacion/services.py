import unicodedata

from dashboard.models import Colaborador, Procedimiento


class ImportacionValidationError(Exception):
    pass


def normalizar(texto):
    texto = str(texto).upper().strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def cargar_workbook(archivo):
    try:
        import openpyxl
        return openpyxl.load_workbook(archivo)
    except Exception as exc:
        raise ImportacionValidationError(
            "No se pudo leer el archivo. Asegúrate de subir un .xlsx válido."
        ) from exc


def procesar_colaboradores_ops(archivo):
    wb = cargar_workbook(archivo)
    hoja = wb["OPS 2026"] if "OPS 2026" in wb.sheetnames else wb.active

    headers = {}
    for col in range(1, hoja.max_column + 1):
        val = hoja.cell(1, col).value
        if val:
            headers[val] = col

    campos_requeridos = ["NOMBRES CONTRATISTA", "APELLIDOS CONTRATISTA"]
    for campo in campos_requeridos:
        if campo not in headers:
            raise ImportacionValidationError(f"No se encontró la columna '{campo}'.")

    colaboradores_bd = {
        normalizar(colaborador.nombre): colaborador
        for colaborador in Colaborador.objects.all()
    }

    creadas, actualizadas, omitidas = [], [], []

    for row in range(2, hoja.max_row + 1):
        dep = (
            hoja.cell(row, headers.get("DEPENDENCIA ASOCIADA", 0)).value
            if "DEPENDENCIA ASOCIADA" in headers
            else None
        )
        if dep and "RED NACIONAL" not in str(dep).upper():
            continue

        nombres = (hoja.cell(row, headers["NOMBRES CONTRATISTA"]).value or "").strip()
        apellidos = (hoja.cell(row, headers["APELLIDOS CONTRATISTA"]).value or "").strip()
        if not nombres and not apellidos:
            continue

        nombre_completo = f"{nombres} {apellidos}".strip()
        norm = normalizar(nombre_completo)

        campos = _campos_colaborador_desde_hoja(hoja, headers, row)

        if norm in colaboradores_bd:
            colaborador = colaboradores_bd[norm]
            for key, value in campos.items():
                if value is not None:
                    setattr(colaborador, key, value)
            colaborador.save()
            actualizadas.append(colaborador.nombre)
        else:
            colaborador = Colaborador.objects.create(
                nombre=nombre_completo,
                **{key: value for key, value in campos.items()},
            )
            colaboradores_bd[norm] = colaborador
            creadas.append(nombre_completo)

    return {
        "creadas": creadas,
        "actualizadas": actualizadas,
        "omitidas": omitidas,
        "total_creadas": len(creadas),
        "total_actualizadas": len(actualizadas),
    }


def _celda_opcional(hoja, headers, row, nombre):
    return hoja.cell(row, headers[nombre]).value if nombre in headers else None


def _normalizar_fecha(value):
    return value.date() if hasattr(value, "date") else value


def _campos_colaborador_desde_hoja(hoja, headers, row):
    cedula = _celda_opcional(hoja, headers, row, "NUMERO DE CEDULA")
    fecha_inicio = _normalizar_fecha(
        _celda_opcional(hoja, headers, row, "FECHA ESTIMADA INICIO CONTRATO")
    )
    fecha_fin = _normalizar_fecha(
        _celda_opcional(hoja, headers, row, "FECHA TERMINACION CONTRATO")
    )
    honorarios = _celda_opcional(hoja, headers, row, "VALOR HONORARIOS MENSUALES ESTIMADOS")
    objeto = _celda_opcional(hoja, headers, row, "OBJETO PAA")
    proc_texto = _celda_opcional(hoja, headers, row, "PROCEDIMIENTO")

    procedimiento = None
    if proc_texto:
        proc_texto = str(proc_texto).strip().upper()
        procedimiento, _ = Procedimiento.objects.get_or_create(nombre=proc_texto)

    return {
        "cedula": str(cedula) if cedula else None,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "honorarios": honorarios,
        "objeto": str(objeto).strip() if objeto else None,
        "procedimiento": procedimiento,
    }
