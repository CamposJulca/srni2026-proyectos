"""
Importa colaboradores desde data/RED.xlsx a los modelos Procedimiento y Colaborador.

Uso:
    python manage.py importar_colaboradores
    python manage.py importar_colaboradores --limpiar   # borra todo antes de importar
    python manage.py importar_colaboradores --archivo data/otro.xlsx
"""
import unicodedata
from pathlib import Path
from django.core.management.base import BaseCommand
from dashboard.models import Procedimiento, Colaborador


MESES_ES = {
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}


def _parsear_fecha(valor):
    """Convierte fechas de Excel (datetime, date o string 'DD-mes-AA') a date."""
    if valor is None:
        return None
    if hasattr(valor, 'date'):
        return valor.date()
    import datetime
    if isinstance(valor, datetime.date):
        return valor
    s = str(valor).strip().lower()
    # Formato: 28-ene.-26 o 28-ene-26
    partes = s.replace('.', '').split('-')
    if len(partes) == 3:
        try:
            dia = int(partes[0])
            mes = MESES_ES.get(partes[1][:3])
            anio = int(partes[2])
            if anio < 100:
                anio += 2000
            if mes:
                return datetime.date(anio, mes, dia)
        except (ValueError, TypeError):
            pass
    return None


def _normalizar(s):
    s = str(s).upper().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


class Command(BaseCommand):
    help = 'Importa colaboradores desde RED.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            default='data/RED.xlsx',
            help='Ruta al archivo Excel (default: data/RED.xlsx)',
        )
        parser.add_argument(
            '--limpiar',
            action='store_true',
            help='Elimina todos los colaboradores y procedimientos antes de importar',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('openpyxl no está instalado. Ejecuta: pip install openpyxl')
            return

        ruta = Path(options['archivo'])
        if not ruta.exists():
            self.stderr.write(f'Archivo no encontrado: {ruta}')
            return

        if options['limpiar']:
            total_col, _ = Colaborador.objects.all().delete()
            total_proc, _ = Procedimiento.objects.all().delete()
            self.stdout.write(f'Limpieza: {total_col} colaboradores y {total_proc} procedimientos eliminados.')

        wb = openpyxl.load_workbook(ruta)
        hoja = wb['OPS 2026'] if 'OPS 2026' in wb.sheetnames else wb.active
        self.stdout.write(f'Leyendo hoja "{hoja.title}" ({hoja.max_row} filas)...')

        # Construir mapa de cabeceras
        headers = {}
        for c in range(1, hoja.max_column + 1):
            val = hoja.cell(1, c).value
            if val:
                headers[str(val).strip()] = c

        creados = 0
        actualizados = 0
        omitidos = 0
        proc_cache = {}

        # Mapa de colaboradores existentes (normalizado → objeto)
        colaboradores_bd = {_normalizar(c.nombre): c for c in Colaborador.objects.all()}

        for row in range(2, hoja.max_row + 1):
            dep = hoja.cell(row, headers.get('DEPENDENCIA ASOCIADA', 0)).value if 'DEPENDENCIA ASOCIADA' in headers else None
            if dep and 'RED NACIONAL' not in str(dep).upper():
                omitidos += 1
                continue

            nombres   = (hoja.cell(row, headers['NOMBRES CONTRATISTA']).value or '').strip() if 'NOMBRES CONTRATISTA' in headers else ''
            apellidos = (hoja.cell(row, headers['APELLIDOS CONTRATISTA']).value or '').strip() if 'APELLIDOS CONTRATISTA' in headers else ''
            if not nombres and not apellidos:
                continue

            nombre_completo = f'{nombres} {apellidos}'.strip()
            norm = _normalizar(nombre_completo)

            # Cedula
            cedula_raw = hoja.cell(row, headers['NUMERO DE CEDULA']).value if 'NUMERO DE CEDULA' in headers else None
            cedula = str(int(cedula_raw)) if cedula_raw and str(cedula_raw).replace('.', '').isdigit() else (str(cedula_raw).strip() if cedula_raw else None)

            # Fechas
            fecha_inicio = _parsear_fecha(hoja.cell(row, headers['FECHA ESTIMADA INICIO CONTRATO']).value if 'FECHA ESTIMADA INICIO CONTRATO' in headers else None)
            fecha_fin    = _parsear_fecha(hoja.cell(row, headers['FECHA TERMINACION CONTRATO']).value if 'FECHA TERMINACION CONTRATO' in headers else None)

            # Honorarios — usar columna 16 (VALOR HONORARIOS MENSUALES ESTIMADOS)
            honorarios = hoja.cell(row, headers['VALOR HONORARIOS MENSUALES ESTIMADOS']).value if 'VALOR HONORARIOS MENSUALES ESTIMADOS' in headers else None
            if honorarios and not isinstance(honorarios, (int, float)):
                try:
                    honorarios = float(str(honorarios).replace(',', '').replace('$', '').strip())
                except ValueError:
                    honorarios = None

            # Objeto
            objeto = hoja.cell(row, headers['OBJETO PAA']).value if 'OBJETO PAA' in headers else None
            objeto = str(objeto).strip() if objeto else None

            # Procedimiento → FK
            proc_texto = hoja.cell(row, headers['PROCEDIMIENTO']).value if 'PROCEDIMIENTO' in headers else None
            procedimiento_obj = None
            if proc_texto:
                proc_key = str(proc_texto).strip().upper()
                if proc_key not in proc_cache:
                    proc_cache[proc_key], _ = Procedimiento.objects.get_or_create(nombre=proc_key)
                procedimiento_obj = proc_cache[proc_key]

            campos = {
                'cedula':        cedula,
                'procedimiento': procedimiento_obj,
                'fecha_inicio':  fecha_inicio,
                'fecha_fin':     fecha_fin,
                'honorarios':    honorarios,
                'objeto':        objeto,
            }

            if norm in colaboradores_bd:
                col = colaboradores_bd[norm]
                for k, v in campos.items():
                    if v is not None:
                        setattr(col, k, v)
                col.save()
                actualizados += 1
            else:
                col = Colaborador.objects.create(nombre=nombre_completo, **campos)
                colaboradores_bd[norm] = col
                creados += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nImportación completada:'
            f'\n  Procedimientos creados : {len(proc_cache)}'
            f'\n  Colaboradores creados  : {creados}'
            f'\n  Colaboradores actualizados: {actualizados}'
            f'\n  Filas omitidas (otra dependencia): {omitidos}'
        ))
        self.stdout.write('\nProcedimientos encontrados:')
        for p in Procedimiento.objects.all().order_by('nombre'):
            total = Colaborador.objects.filter(procedimiento=p).count()
            self.stdout.write(f'  {p.nombre}: {total} colaboradores')
