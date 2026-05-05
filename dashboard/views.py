from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Avg, Count, Q, Sum
from datetime import date
from calendar import monthrange
from django.db import connection
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import json
import re
import unicodedata

from .models import (
    Procedimiento,
    Proyecto,
    Modulo,
    Colaborador,
    Rol,
    Asignacion,
    Obligacion,
    Actividad,
    CuentaCobro,
    Perfil,
    EvidenciaActividad,
    ReporteSemanal,
)
from .permisos import admin_required, _es_admin


# ============================================================
#  CONFIGURACIÓN CRUD
# ============================================================

PROCEDIMIENTOS = [
    "INSTRUMENTALIZACIÓN",
    "EQUIPO BASE",
    "ANÁLISIS",
    "CARACTERIZACIÓN",
    "DIFUSIÓN Y APRENDIZAJE",
    "AIDI",
    "MESA DE SERVICIOS",
    "GIS",
]

def login_view(request):
    if request.user.is_authenticated:
        return redirect("/")
    error = None
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            next_url = request.GET.get("next", "")
            if next_url:
                return redirect(next_url)
            perfil = getattr(user, 'perfil', None)
            if perfil and perfil.rol == 'colaborador':
                return redirect('/mi-cronograma/')
            return redirect('/')
        error = "Usuario o contraseña incorrectos."
    return render(request, "dashboard/login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("/login/")


@login_required
def home(request):
    perfil = getattr(request.user, 'perfil', None)
    if perfil and perfil.rol == 'colaborador':
        return redirect('/mi-cronograma/')
    return render(request, "dashboard/home.html")


@admin_required
def consultas_view(request):
    return render(request, "dashboard/consultas_view.html", {
        "modulo_activo": "consultas",
    })


@admin_required
def crud_main_view(request):
    return render(request, "dashboard/crud_view.html", {
        "modulo_activo": "crud",
        "procedimientos": PROCEDIMIENTOS,
    })


@admin_required
def gerencial_view(request):
    return render(request, "dashboard/gerencial_view.html", {
        "modulo_activo": "gerencial",
    })


@admin_required
def gerencial_data(request):
    hoy = date.today()

    total       = Colaborador.objects.count()
    masa        = Colaborador.objects.filter(honorarios__isnull=False).aggregate(t=Sum('honorarios'))['t'] or 0
    vigentes    = Colaborador.objects.filter(fecha_fin__isnull=False, fecha_fin__gte=hoy).count()
    vencidos    = Colaborador.objects.filter(fecha_fin__isnull=False, fecha_fin__lt=hoy).count()
    sin_fecha   = Colaborador.objects.filter(fecha_fin__isnull=True).count()

    por_proc_qs = list(
        Colaborador.objects
        .values('procedimiento__nombre')
        .annotate(personas=Count('id'), masa=Sum('honorarios'))
        .order_by('-personas')
    )
    por_proc = []
    for p in por_proc_qs:
        por_proc.append({
            'procedimiento': p['procedimiento__nombre'] or 'Sin clasificar',
            'personas': p['personas'],
            'masa': float(p['masa'] or 0),
        })

    return JsonResponse({
        'kpis': {
            'total':     total,
            'masa':      float(masa),
            'vigentes':  vigentes,
            'vencidos':  vencidos,
            'sin_fecha': sin_fecha,
        },
        'por_procedimiento': por_proc,
        'compromisos_mes':   [],
    })


@admin_required
def dashboard(request):

    procedimiento = request.GET.get("procedimiento", "INSTRUMENTALIZACIÓN")

    if procedimiento:
        colaboradores = Colaborador.objects.filter(
            procedimiento__nombre=procedimiento
        ).order_by("nombre")
    else:
        colaboradores = Colaborador.objects.all().order_by("nombre")

    roles = Rol.objects.all()
    proyectos = Proyecto.objects.all()

    return render(
        request,
        "dashboard/dashboard_view.html",
        {
            "personas": colaboradores,
            "roles": roles,
            "proyectos": proyectos,
            "procedimientos": PROCEDIMIENTOS,
            "procedimiento_activo": procedimiento,
            "modulo_activo": "dashboard",
        }
    )


def _resolver_periodo(periodo_texto):
    if not periodo_texto:
        hoy = date.today()
        return date(hoy.year, hoy.month, 1)

    m = re.match(r"^(\d{4})-(\d{2})$", periodo_texto)
    if not m:
        return None

    year = int(m.group(1))
    month = int(m.group(2))
    if not 1 <= month <= 12:
        return None

    return date(year, month, 1)


def _clasificar_cruce_cobro(cuenta, cumplimiento_pct, cobro_pct):
    if not cuenta:
        return "sin_cuenta"
    if cuenta.estado == "rechazada":
        return "rechazada"
    if cobro_pct <= (cumplimiento_pct + 5):
        return "alineado"
    if cobro_pct <= (cumplimiento_pct + 20):
        return "alerta"
    return "desalineado"


@admin_required
def dashboard_data(request):

    persona = request.GET.get("persona")
    rol = request.GET.get("rol")
    proyecto = request.GET.get("proyecto")
    procedimiento = request.GET.get("procedimiento")
    periodo_texto = request.GET.get("periodo", "").strip()

    periodo = _resolver_periodo(periodo_texto)
    if periodo is None:
        return JsonResponse(
            {"error": "Parámetro 'periodo' inválido. Use formato YYYY-MM."},
            status=400,
        )
    ultimo_dia = monthrange(periodo.year, periodo.month)[1]
    periodo_fin = date(periodo.year, periodo.month, ultimo_dia)

    asignaciones = Asignacion.objects.select_related(
        "colaborador",
        "rol",
        "modulo",
        "modulo__proyecto"
    )

    if procedimiento:
        asignaciones = asignaciones.filter(colaborador__procedimiento__nombre=procedimiento)
    if persona:
        asignaciones = asignaciones.filter(colaborador_id=persona)
    if rol:
        asignaciones = asignaciones.filter(rol_id=rol)
    if proyecto:
        asignaciones = asignaciones.filter(modulo__proyecto_id=proyecto)

    # ==============================
    # KPIs
    # ==============================
    colaboradores_qs = Colaborador.objects.all()
    if procedimiento:
        colaboradores_qs = colaboradores_qs.filter(procedimiento__nombre=procedimiento)
    if persona:
        colaboradores_qs = colaboradores_qs.filter(pk=persona)
    if rol:
        colaboradores_qs = colaboradores_qs.filter(asignaciones__rol_id=rol)
    if proyecto:
        colaboradores_qs = colaboradores_qs.filter(asignaciones__modulo__proyecto_id=proyecto)
    colaboradores_qs = colaboradores_qs.distinct()

    kpis = {
        "proyectos": Proyecto.objects.count(),
        "modulos": Modulo.objects.count(),
        "personas": colaboradores_qs.count(),
        "asignaciones": asignaciones.count(),
    }

    # ==============================
    # Proyectos por colaborador
    # ==============================
    proyectos_persona = (
        asignaciones
        .values("colaborador__nombre")
        .annotate(total=Count("modulo__proyecto", distinct=True))
    )

    # ==============================
    # Roles por colaborador
    # ==============================
    roles_persona = (
        asignaciones
        .values("rol__nombre")
        .annotate(total=Count("colaborador", distinct=True))
    )

    # ==============================
    # Módulos filtrados
    # ==============================
    modulos_ids = asignaciones.values_list("modulo_id", flat=True)
    modulos_proyecto = (
        Modulo.objects
        .filter(id__in=modulos_ids)
        .values("proyecto__nombre")
        .annotate(total=Count("id", distinct=True))
    )

    # ==============================
    # Cruce cuentas de cobro vs avance/cumplimiento (mensual)
    # ==============================
    colaboradores = list(colaboradores_qs.order_by("nombre"))
    colaboradores_ids = [c.id for c in colaboradores]

    cuentas_qs = CuentaCobro.objects.filter(
        colaborador_id__in=colaboradores_ids,
        periodo=periodo,
    )
    cuentas_map = {c.colaborador_id: c for c in cuentas_qs}

    actividades_qs = Actividad.objects.filter(
        obligacion__colaborador_id__in=colaboradores_ids,
        fecha_inicio__lte=periodo_fin,
        fecha_fin__gte=periodo,
    )
    if proyecto:
        actividades_qs = actividades_qs.filter(proyecto_id=proyecto)

    actividad_por_colab = {
        row["obligacion__colaborador_id"]: row
        for row in actividades_qs.values("obligacion__colaborador_id").annotate(
            total=Count("id"),
            completadas=Count("id", filter=Q(estado="completada")),
            avance=Avg("progreso"),
        )
    }

    cruce_cobro = []
    total_actividades_periodo = 0
    total_completadas_periodo = 0
    suma_avance_ponderado = 0.0

    for colaborador in colaboradores:
        resumen = actividad_por_colab.get(colaborador.id, {})
        total = int(resumen.get("total") or 0)
        completadas = int(resumen.get("completadas") or 0)
        avance = float(resumen.get("avance") or 0.0)

        cumplimiento = round((completadas * 100 / total), 1) if total else 0.0
        avance = round(avance, 1)

        cuenta = cuentas_map.get(colaborador.id)
        valor_cobrado = float(cuenta.valor_cobrado) if cuenta else 0.0
        honorarios = float(colaborador.honorarios or 0.0)
        cobro_pct = round((valor_cobrado * 100 / honorarios), 1) if honorarios > 0 else 0.0

        total_actividades_periodo += total
        total_completadas_periodo += completadas
        suma_avance_ponderado += (avance * total)

        cruce_cobro.append({
            "colaborador_id": colaborador.id,
            "colaborador": colaborador.nombre,
            "periodo": periodo.isoformat(),
            "actividades_periodo": total,
            "actividades_completadas": completadas,
            "avance_pct": avance,
            "cumplimiento_pct": cumplimiento,
            "honorarios_mensuales": honorarios,
            "valor_cobrado": valor_cobrado,
            "cobro_pct": cobro_pct,
            "brecha_cumplimiento_cobro": round(cumplimiento - cobro_pct, 1),
            "estado_cuenta": cuenta.estado if cuenta else "sin_registro",
            "numero_cuenta": cuenta.numero_cuenta if cuenta else None,
            "estado_cruce": _clasificar_cruce_cobro(cuenta, cumplimiento, cobro_pct),
        })

    avance_promedio_periodo = (
        round(suma_avance_ponderado / total_actividades_periodo, 1)
        if total_actividades_periodo else 0.0
    )
    cumplimiento_periodo = (
        round(total_completadas_periodo * 100 / total_actividades_periodo, 1)
        if total_actividades_periodo else 0.0
    )

    kpis["cuentas_periodo"] = len(cuentas_map)
    kpis["avance_periodo"] = avance_promedio_periodo
    kpis["cumplimiento_periodo"] = cumplimiento_periodo

    data = {
        "kpis": kpis,
        "proyectos_persona": list(proyectos_persona),
        "roles_persona": list(roles_persona),
        "modulos_proyecto": list(modulos_proyecto),
        "compromisos_mes": [],
        "periodo_cobro": periodo.strftime("%Y-%m"),
        "cruce_cobro": cruce_cobro,
    }

    return JsonResponse(data)


@admin_required
@require_POST
def sql_query(request):
    try:
        body = json.loads(request.body)
        query = body.get("query", "").strip()
    except Exception:
        return JsonResponse({"error": "JSON inválido."}, status=400)

    if not query:
        return JsonResponse({"error": "La consulta está vacía."}, status=400)

    # Solo permite SELECT
    primera_palabra = re.split(r'\s+', query)[0].upper()
    if primera_palabra != "SELECT":
        return JsonResponse(
            {"error": "Solo se permiten consultas SELECT."},
            status=403
        )

    # Bloquea palabras peligrosas por si cuelan en subconsultas
    peligrosas = r'\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|TRUNCATE|REPLACE|ATTACH|DETACH|PRAGMA)\b'
    if re.search(peligrosas, query, re.IGNORECASE):
        return JsonResponse(
            {"error": "La consulta contiene operaciones no permitidas."},
            status=403
        )

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            columnas = [desc[0] for desc in cursor.description]
            filas = cursor.fetchmany(500)  # máximo 500 filas
        return JsonResponse({
            "columnas": columnas,
            "filas": [list(f) for f in filas],
            "total": len(filas)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ============================================================
#  CRUD API
# ============================================================

def _opciones_fk(modelo, campo_display="nombre"):
    try:
        qs = modelo.objects.all().order_by(campo_display)
        return [{"id": o.pk, "label": str(getattr(o, campo_display, None) or o)} for o in qs]
    except Exception:
        return [{"id": o.pk, "label": str(o)} for o in modelo.objects.all()]


TABLA_META = {
    "colaborador": {
        "modelo": Colaborador,
        "label": "Colaboradores",
        "buscar_en": ["nombre", "cedula"],
        "filtro_colaborador": False,
        "campos": [
            {"name": "nombre",           "label": "Nombre completo", "type": "text",     "required": True},
            {"name": "cedula",           "label": "Cédula",          "type": "text"},
            {"name": "procedimiento_id", "label": "Procedimiento",   "type": "fk",       "fk_modelo": "Procedimiento"},
            {"name": "fecha_inicio",     "label": "Fecha inicio",    "type": "date"},
            {"name": "fecha_fin",        "label": "Fecha fin",       "type": "date"},
            {"name": "honorarios",       "label": "Honorarios",      "type": "number"},
            {"name": "objeto",           "label": "Objeto contrato", "type": "textarea"},
        ],
        "columnas_lista": ["nombre", "cedula", "procedimiento__nombre", "fecha_inicio", "fecha_fin", "honorarios"],
    },
    "obligacion": {
        "modelo": Obligacion,
        "label": "Obligaciones",
        "buscar_en": ["descripcion"],
        "filtro_colaborador": True,
        "campos": [
            {"name": "colaborador_id", "label": "Colaborador", "type": "fk",      "required": True, "fk_modelo": "Colaborador"},
            {"name": "descripcion",    "label": "Descripción", "type": "textarea","required": True},
        ],
        "columnas_lista": ["colaborador__nombre", "descripcion"],
    },
    "actividad": {
        "modelo": Actividad,
        "label": "Actividades",
        "buscar_en": ["descripcion", "actividad_id"],
        "filtro_colaborador": True,
        "campos": [
            {"name": "obligacion_id",  "label": "Obligación",   "type": "fk",      "required": True, "fk_modelo": "Obligacion"},
            {"name": "actividad_id",   "label": "ID actividad", "type": "text"},
            {"name": "descripcion",    "label": "Descripción",  "type": "textarea","required": True},
            {"name": "fecha_inicio",   "label": "Fecha inicio", "type": "date"},
            {"name": "fecha_fin",      "label": "Fecha fin",    "type": "date"},
            {"name": "progreso",       "label": "Progreso (%)", "type": "number"},
            {"name": "estado",         "label": "Estado",       "type": "choice",
             "opciones_fijas": [
                {"id": "pendiente",   "label": "Pendiente"},
                {"id": "en_curso",    "label": "En curso"},
                {"id": "completada",  "label": "Completada"},
             ]},
            {"name": "orden",          "label": "Orden",        "type": "number"},
        ],
        "columnas_lista": ["obligacion__colaborador__nombre", "actividad_id", "descripcion", "estado", "progreso"],
    },
    "cuenta_cobro": {
        "modelo": CuentaCobro,
        "label": "Cuentas de cobro",
        "buscar_en": ["numero_cuenta", "estado", "colaborador__nombre"],
        "filtro_colaborador": True,
        "campos": [
            {"name": "colaborador_id",  "label": "Colaborador", "type": "fk",      "required": True, "fk_modelo": "Colaborador"},
            {"name": "periodo",         "label": "Periodo (día 1 del mes)", "type": "date", "required": True},
            {"name": "numero_cuenta",   "label": "Número cuenta", "type": "text"},
            {"name": "fecha_radicacion","label": "Fecha radicación", "type": "date"},
            {"name": "valor_cobrado",   "label": "Valor cobrado", "type": "number", "required": True},
            {"name": "estado",          "label": "Estado", "type": "choice",
             "opciones_fijas": [
                {"id": "borrador", "label": "Borrador"},
                {"id": "radicada", "label": "Radicada"},
                {"id": "aprobada", "label": "Aprobada"},
                {"id": "pagada", "label": "Pagada"},
                {"id": "rechazada", "label": "Rechazada"},
             ]},
            {"name": "observaciones",   "label": "Observaciones", "type": "textarea"},
        ],
        "columnas_lista": ["colaborador__nombre", "periodo", "numero_cuenta", "valor_cobrado", "estado", "fecha_radicacion"],
    },
    "asignacion": {
        "modelo": Asignacion,
        "label": "Asignaciones",
        "buscar_en": [],
        "filtro_colaborador": False,
        "campos": [
            {"name": "colaborador_id", "label": "Colaborador", "type": "fk", "required": True, "fk_modelo": "Colaborador"},
            {"name": "rol_id",         "label": "Rol",         "type": "fk", "required": True, "fk_modelo": "Rol"},
            {"name": "modulo_id",      "label": "Módulo",      "type": "fk", "required": True, "fk_modelo": "Modulo"},
        ],
        "columnas_lista": ["colaborador__nombre", "rol__nombre", "modulo__nombre", "modulo__proyecto__nombre"],
    },
    "modulo": {
        "modelo": Modulo,
        "label": "Módulos",
        "buscar_en": ["nombre", "referente"],
        "filtro_colaborador": False,
        "campos": [
            {"name": "proyecto_id", "label": "Proyecto",  "type": "fk",  "required": True, "fk_modelo": "Proyecto"},
            {"name": "nombre",      "label": "Nombre",    "type": "text", "required": True},
            {"name": "referente",   "label": "Referente", "type": "text"},
        ],
        "columnas_lista": ["nombre", "proyecto__nombre", "referente"],
    },
    "rol": {
        "modelo": Rol,
        "label": "Roles",
        "buscar_en": ["nombre"],
        "filtro_colaborador": False,
        "campos": [
            {"name": "nombre", "label": "Nombre", "type": "text", "required": True},
        ],
        "columnas_lista": ["nombre"],
    },
}

FK_MODELOS = {
    "Procedimiento": Procedimiento,
    "Colaborador":   Colaborador,
    "Proyecto":      Proyecto,
    "Modulo":        Modulo,
    "Rol":           Rol,
    "Obligacion":    Obligacion,
    "Actividad":     Actividad,
    "CuentaCobro":   CuentaCobro,
}


def _serializar_fila(obj, columnas):
    fila = {"id": obj.pk}
    for col in columnas:
        if "__" in col:
            partes = col.split("__")
            val = obj
            for p in partes:
                val = getattr(val, p, None)
                if val is None:
                    break
        else:
            val = getattr(obj, col, None)
        if hasattr(val, 'isoformat'):
            val = val.isoformat()
        fila[col] = str(val) if val is not None else None
    return fila


@admin_required
def crud_meta(request):
    """Devuelve metadatos de todas las tablas (campos, conteos, opciones FK)."""
    resultado = {}
    for key, cfg in TABLA_META.items():
        campos = []
        for c in cfg["campos"]:
            campo = dict(c)
            if campo["type"] == "fk":
                fk_m = FK_MODELOS.get(campo.get("fk_modelo"))
                campo["opciones"] = _opciones_fk(fk_m) if fk_m else []
            elif campo["type"] == "choice":
                campo["opciones"] = campo.get("opciones_fijas", [])
            campos.append(campo)
        resultado[key] = {
            "label":              cfg["label"],
            "campos":             campos,
            "columnas":           cfg["columnas_lista"],
            "filtro_colaborador": cfg.get("filtro_colaborador", False),
            "total":              cfg["modelo"].objects.count(),
        }
    return JsonResponse(resultado)


@admin_required
def crud_lista(request, tabla):
    cfg = TABLA_META.get(tabla)
    if not cfg:
        return JsonResponse({"error": "Tabla no válida."}, status=404)

    qs = cfg["modelo"].objects.all()

    # Filtro por procedimiento (colaborador)
    proc = request.GET.get("proc", "").strip()
    if proc and tabla == "colaborador":
        qs = qs.filter(procedimiento__nombre=proc)

    # Filtro por colaborador (obligacion / actividad)
    colab_id = request.GET.get("colaborador", "").strip()
    if colab_id:
        if tabla == "obligacion":
            qs = qs.filter(colaborador_id=colab_id)
        elif tabla == "actividad":
            qs = qs.filter(obligacion__colaborador_id=colab_id)
        elif tabla == "cuenta_cobro":
            qs = qs.filter(colaborador_id=colab_id)

    # Búsqueda
    q = request.GET.get("q", "").strip()
    if q and cfg["buscar_en"]:
        filtro = Q()
        for campo in cfg["buscar_en"]:
            filtro |= Q(**{f"{campo}__icontains": q})
        qs = qs.filter(filtro)

    # Relacionados
    if tabla == "modulo":
        qs = qs.select_related("proyecto")
    elif tabla == "asignacion":
        qs = qs.select_related("colaborador", "rol", "modulo", "modulo__proyecto")
    elif tabla == "colaborador":
        qs = qs.select_related("procedimiento")
    elif tabla == "obligacion":
        qs = qs.select_related("colaborador")
    elif tabla == "actividad":
        qs = qs.select_related("obligacion", "obligacion__colaborador")
    elif tabla == "cuenta_cobro":
        qs = qs.select_related("colaborador")

    total = qs.count()
    page  = int(request.GET.get("page", 1))
    size  = 25
    qs    = qs[(page-1)*size : page*size]

    filas = [_serializar_fila(obj, cfg["columnas_lista"]) for obj in qs]
    return JsonResponse({"filas": filas, "total": total, "page": page, "size": size})


@admin_required
def crud_detalle(request, tabla, pk):
    cfg = TABLA_META.get(tabla)
    if not cfg:
        return JsonResponse({"error": "Tabla no válida."}, status=404)

    try:
        obj = cfg["modelo"].objects.get(pk=pk)
    except cfg["modelo"].DoesNotExist:
        return JsonResponse({"error": "Registro no encontrado."}, status=404)

    if request.method == "GET":
        datos = {"id": obj.pk}
        for c in cfg["campos"]:
            val = getattr(obj, c["name"], None)
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            datos[c["name"]] = val
        return JsonResponse(datos)

    if request.method in ("PUT", "PATCH"):
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({"error": "JSON inválido."}, status=400)
        for c in cfg["campos"]:
            nombre = c["name"]
            if nombre in body:
                val = body[nombre] or None
                setattr(obj, nombre, val)
        try:
            obj.save()
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        return JsonResponse({"ok": True, "id": obj.pk})

    if request.method == "DELETE":
        try:
            obj.delete()
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        return JsonResponse({"ok": True})

    return JsonResponse({"error": "Método no permitido."}, status=405)


@admin_required
def crud_crear(request, tabla):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido."}, status=405)

    cfg = TABLA_META.get(tabla)
    if not cfg:
        return JsonResponse({"error": "Tabla no válida."}, status=404)

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "JSON inválido."}, status=400)

    kwargs = {}
    for c in cfg["campos"]:
        nombre = c["name"]
        val = body.get(nombre) or None
        kwargs[nombre] = val

    try:
        obj = cfg["modelo"].objects.create(**kwargs)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"ok": True, "id": obj.pk}, status=201)


def _normalizar(s):
    s = str(s).upper().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def personas_por_rol(request):
    rol_nombre = request.GET.get("rol", "")
    proyecto_id = request.GET.get("proyecto", "")

    qs = Asignacion.objects.select_related(
        "colaborador", "rol", "modulo", "modulo__proyecto"
    )

    if rol_nombre:
        qs = qs.filter(rol__nombre=rol_nombre)
    if proyecto_id:
        qs = qs.filter(modulo__proyecto_id=proyecto_id)

    data = [
        {
            "persona":  a.colaborador.nombre,
            "modulo":   a.modulo.nombre,
            "proyecto": a.modulo.proyecto.nombre,
        }
        for a in qs.order_by("colaborador__nombre")
    ]

    return JsonResponse({"asignaciones": data})


@admin_required
def carga(request):
    if request.method == "GET":
        return render(request, "dashboard/carga.html")

    # POST — procesar archivo
    archivo = request.FILES.get("archivo")
    if not archivo:
        return JsonResponse({"error": "No se recibió ningún archivo."}, status=400)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(archivo)
    except Exception:
        return JsonResponse({"error": "No se pudo leer el archivo. Asegúrate de subir un .xlsx válido."}, status=400)

    if "OPS 2026" not in wb.sheetnames:
        hoja = wb.active
    else:
        hoja = wb["OPS 2026"]

    # Construir mapa de cabeceras
    headers = {}
    for c in range(1, hoja.max_column + 1):
        val = hoja.cell(1, c).value
        if val:
            headers[val] = c

    campos_requeridos = ["NOMBRES CONTRATISTA", "APELLIDOS CONTRATISTA"]
    for campo in campos_requeridos:
        if campo not in headers:
            return JsonResponse({"error": f"No se encontró la columna '{campo}'."}, status=400)

    # Construir mapa de colaboradores existentes en BD (normalizado → objeto)
    colaboradores_bd = {_normalizar(c.nombre): c for c in Colaborador.objects.all()}

    creadas, actualizadas, omitidas = [], [], []

    for row in range(2, hoja.max_row + 1):
        dep = hoja.cell(row, headers.get("DEPENDENCIA ASOCIADA", 0)).value if "DEPENDENCIA ASOCIADA" in headers else None
        if dep and "RED NACIONAL" not in str(dep).upper():
            continue

        nombres = (hoja.cell(row, headers["NOMBRES CONTRATISTA"]).value or "").strip()
        apellidos = (hoja.cell(row, headers["APELLIDOS CONTRATISTA"]).value or "").strip()
        if not nombres and not apellidos:
            continue

        nombre_completo = f"{nombres} {apellidos}".strip()
        norm = _normalizar(nombre_completo)

        # Extraer campos extra
        cedula = hoja.cell(row, headers["NUMERO DE CEDULA"]).value if "NUMERO DE CEDULA" in headers else None
        fecha_inicio = hoja.cell(row, headers["FECHA ESTIMADA INICIO CONTRATO"]).value if "FECHA ESTIMADA INICIO CONTRATO" in headers else None
        fecha_fin = hoja.cell(row, headers["FECHA TERMINACION CONTRATO"]).value if "FECHA TERMINACION CONTRATO" in headers else None
        honorarios = hoja.cell(row, headers["VALOR HONORARIOS MENSUALES ESTIMADOS"]).value if "VALOR HONORARIOS MENSUALES ESTIMADOS" in headers else None
        objeto = hoja.cell(row, headers["OBJETO PAA"]).value if "OBJETO PAA" in headers else None
        proc_texto = hoja.cell(row, headers["PROCEDIMIENTO"]).value if "PROCEDIMIENTO" in headers else None

        # Normalizar fechas
        if hasattr(fecha_inicio, 'date'):
            fecha_inicio = fecha_inicio.date()
        if hasattr(fecha_fin, 'date'):
            fecha_fin = fecha_fin.date()

        # Resolver FK de procedimiento
        procedimiento_obj = None
        if proc_texto:
            proc_texto = str(proc_texto).strip().upper()
            procedimiento_obj, _ = Procedimiento.objects.get_or_create(nombre=proc_texto)

        campos = {
            "cedula":       str(cedula) if cedula else None,
            "fecha_inicio": fecha_inicio,
            "fecha_fin":    fecha_fin,
            "honorarios":   honorarios,
            "objeto":       str(objeto).strip() if objeto else None,
            "procedimiento": procedimiento_obj,
        }

        if norm in colaboradores_bd:
            colaborador = colaboradores_bd[norm]
            for k, v in campos.items():
                if v is not None:
                    setattr(colaborador, k, v)
            colaborador.save()
            actualizadas.append(colaborador.nombre)
        else:
            colaborador = Colaborador.objects.create(nombre=nombre_completo, **{k: v for k, v in campos.items()})
            colaboradores_bd[norm] = colaborador
            creadas.append(nombre_completo)

    return JsonResponse({
        "creadas": creadas,
        "actualizadas": actualizadas,
        "omitidas": omitidas,
        "total_creadas": len(creadas),
        "total_actualizadas": len(actualizadas),
    })

# ============================================================
#  MÓDULO ACTIVIDADES — CRONOGRAMA GANTT
# ============================================================

@login_required
def actividades_view(request):
    # Todos los colaboradores (con y sin cronograma)
    colaboradores = sorted(
        Colaborador.objects.values_list('nombre', flat=True).distinct()
    )

    obligaciones = list(
        Actividad.objects
        .values_list('obligacion__descripcion', flat=True)
        .distinct()
        .order_by('obligacion__descripcion')
    )
    proyectos = list(
        Proyecto.objects.values_list('nombre', flat=True).order_by('nombre')
    )
    procedimientos = list(
        Procedimiento.objects.values_list('nombre', flat=True).order_by('nombre')
    )
    # Colaboradores agrupados por procedimiento para el JS
    colabs_por_proc = {}
    for c in Colaborador.objects.select_related('procedimiento').order_by('nombre'):
        proc = c.procedimiento.nombre if c.procedimiento else 'Sin procedimiento'
        colabs_por_proc.setdefault(proc, []).append(c.nombre)

    return render(request, 'dashboard/actividades.html', {
        'modulo_activo': 'actividades',
        'colaboradores': colaboradores,
        'obligaciones': obligaciones,
        'proyectos': proyectos,
        'procedimientos': procedimientos,
        'colabs_por_proc_json': json.dumps(colabs_por_proc),
    })


@login_required
def actividades_data(request):
    qs = Actividad.objects.select_related('obligacion__colaborador', 'proyecto').all()
    colaborador = request.GET.get('colaborador', '').strip()
    estado = request.GET.get('estado', '').strip()
    obligacion = request.GET.get('obligacion', '').strip()
    proyecto = request.GET.get('proyecto', '').strip()

    if colaborador:
        qs = qs.filter(obligacion__colaborador__nombre=colaborador)
    if estado:
        qs = qs.filter(estado=estado)
    if obligacion:
        qs = qs.filter(obligacion__descripcion=obligacion)
    if proyecto:
        # Filtrar por colaboradores asignados a ese proyecto
        colab_ids = Asignacion.objects.filter(
            modulo__proyecto__nombre=proyecto
        ).values_list('colaborador_id', flat=True).distinct()
        qs = qs.filter(obligacion__colaborador_id__in=colab_ids)

    # Pre-cargar proyectos por colaborador para la respuesta
    colab_proyectos = {}
    for asig in Asignacion.objects.select_related('modulo__proyecto').all():
        cid = asig.colaborador_id
        pname = asig.modulo.proyecto.nombre
        colab_proyectos.setdefault(cid, set()).add(pname)

    hoy = date.today()
    tasks = []
    for a in qs:
        estado_visual = a.estado
        if a.estado == 'pendiente':
            if a.fecha_fin < hoy:
                estado_visual = 'vencida'
            elif a.fecha_inicio <= hoy <= a.fecha_fin:
                estado_visual = 'en_curso'

        cid = a.obligacion.colaborador_id
        proyectos_colab = sorted(colab_proyectos.get(cid, set()))

        tasks.append({
            'id':           a.pk,
            'colaborador':  a.obligacion.colaborador.nombre,
            'obligacion':   a.obligacion.descripcion,
            'actividad_id': a.actividad_id,
            'descripcion':  a.descripcion,
            'fecha_inicio': a.fecha_inicio.isoformat(),
            'fecha_fin':    a.fecha_fin.isoformat(),
            'progreso':     a.progreso,
            'estado':       a.estado,
            'estado_visual': estado_visual,
            'orden':        a.orden,
            'proyecto':     a.proyecto.nombre if a.proyecto else (proyectos_colab[0] if len(proyectos_colab) == 1 else None),
            'proyectos':    proyectos_colab,
        })

    return JsonResponse({'tasks': tasks})


@admin_required
@require_POST
def actividad_crear(request):
    try:
        data = json.loads(request.body)

        nombre_colab = data.get('colaborador', '').strip()
        desc_oblig   = data.get('obligacion', '').strip()

        if not nombre_colab:
            return JsonResponse({'ok': False, 'error': 'colaborador requerido'}, status=400)

        colaborador_obj = Colaborador.objects.filter(nombre=nombre_colab).first()
        if not colaborador_obj:
            return JsonResponse({'ok': False, 'error': f'Colaborador "{nombre_colab}" no encontrado'}, status=400)

        obligacion_obj, _ = Obligacion.objects.get_or_create(
            colaborador=colaborador_obj,
            descripcion=desc_oblig,
        )

        proyecto_id = data.get('proyecto_id')
        proyecto_obj = Proyecto.objects.filter(pk=proyecto_id).first() if proyecto_id else None

        a = Actividad.objects.create(
            obligacion=obligacion_obj,
            proyecto=proyecto_obj,
            actividad_id=data.get('actividad_id', ''),
            descripcion=data.get('descripcion', ''),
            fecha_inicio=data['fecha_inicio'],
            fecha_fin=data['fecha_fin'],
            progreso=int(data.get('progreso', 0)),
            estado=data.get('estado', 'pendiente'),
        )
        return JsonResponse({'ok': True, 'id': a.pk})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
def actividad_detalle(request, pk):
    try:
        a = Actividad.objects.select_related('obligacion__colaborador', 'proyecto').get(pk=pk)
    except Actividad.DoesNotExist:
        return JsonResponse({'error': 'No encontrada'}, status=404)

    if request.method == 'GET':
        return JsonResponse({
            'id':           a.pk,
            'colaborador':  a.obligacion.colaborador.nombre,
            'obligacion':   a.obligacion.descripcion,
            'actividad_id': a.actividad_id,
            'descripcion':  a.descripcion,
            'fecha_inicio': a.fecha_inicio.isoformat(),
            'fecha_fin':    a.fecha_fin.isoformat(),
            'progreso':     a.progreso,
            'estado':       a.estado,
            'evidencia':    a.evidencia or '',
        })

    # PUT/PATCH/DELETE solo para admins
    # _es_admin importado a nivel de módulo
    if not _es_admin(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method in ('PUT', 'PATCH'):
        data = json.loads(request.body)
        # Bloquear cierre sin evidencia
        if data.get('estado') == 'completada' and a.evidencias.count() == 0:
            return JsonResponse({
                'ok': False,
                'error': 'No se puede marcar como completada sin al menos una evidencia adjunta.',
            }, status=400)
        for field in ('actividad_id', 'descripcion', 'fecha_inicio', 'fecha_fin', 'estado'):
            if field in data:
                setattr(a, field, data[field])
        if 'progreso' in data:
            a.progreso = int(data['progreso'])
        if 'evidencia' in data:
            a.evidencia = data['evidencia']
        a.save()
        return JsonResponse({'ok': True})

    if request.method == 'DELETE':
        a.delete()
        return JsonResponse({'ok': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ============================================================
#  VISTA SEMANAL — compromisos por persona/semana
# ============================================================

# Orden cronológico de todas las semanas 2026
SEMANAS_ORDENADAS = [
    'ENE S1','ENE S2','ENE S3','ENE S4',
    'FEB S1','FEB S2','FEB S3','FEB S4',
    'MAR S1','MAR S2','MAR S3','MAR S4',
    'ABR S1','ABR S2','ABR S3','ABR S4',
    'MAY S1','MAY S2','MAY S3','MAY S4',
    'JUN S1','JUN S2','JUN S3','JUN S4',
    'JUL S1','JUL S2','JUL S3','JUL S4',
    'AGO S1','AGO S2','AGO S3','AGO S4',
    'SEP S1','SEP S2','SEP S3','SEP S4',
    'OCT S1','OCT S2','OCT S3','OCT S4',
    'NOV S1','NOV S2','NOV S3','NOV S4',
    'DIC S1','DIC S2','DIC S3','DIC S4',
]

SEMANA_FECHAS = {
    'ENE S1':('2026-01-05','2026-01-09'), 'ENE S2':('2026-01-12','2026-01-16'),
    'ENE S3':('2026-01-19','2026-01-23'), 'ENE S4':('2026-01-26','2026-01-30'),
    'FEB S1':('2026-02-02','2026-02-06'), 'FEB S2':('2026-02-09','2026-02-13'),
    'FEB S3':('2026-02-16','2026-02-20'), 'FEB S4':('2026-02-23','2026-02-27'),
    'MAR S1':('2026-03-02','2026-03-06'), 'MAR S2':('2026-03-09','2026-03-13'),
    'MAR S3':('2026-03-16','2026-03-20'), 'MAR S4':('2026-03-23','2026-03-27'),
    'ABR S1':('2026-04-06','2026-04-10'), 'ABR S2':('2026-04-13','2026-04-17'),
    'ABR S3':('2026-04-20','2026-04-24'), 'ABR S4':('2026-04-27','2026-04-30'),
    'MAY S1':('2026-05-04','2026-05-08'), 'MAY S2':('2026-05-11','2026-05-15'),
    'MAY S3':('2026-05-18','2026-05-22'), 'MAY S4':('2026-05-25','2026-05-29'),
    'JUN S1':('2026-06-01','2026-06-05'), 'JUN S2':('2026-06-08','2026-06-12'),
    'JUN S3':('2026-06-15','2026-06-19'), 'JUN S4':('2026-06-22','2026-06-26'),
    'JUL S1':('2026-07-06','2026-07-10'), 'JUL S2':('2026-07-13','2026-07-17'),
    'JUL S3':('2026-07-20','2026-07-24'), 'JUL S4':('2026-07-27','2026-07-31'),
    'AGO S1':('2026-08-03','2026-08-07'), 'AGO S2':('2026-08-10','2026-08-14'),
    'AGO S3':('2026-08-17','2026-08-21'), 'AGO S4':('2026-08-24','2026-08-28'),
    'SEP S1':('2026-09-07','2026-09-11'), 'SEP S2':('2026-09-14','2026-09-18'),
    'SEP S3':('2026-09-21','2026-09-25'), 'SEP S4':('2026-09-28','2026-10-02'),
    'OCT S1':('2026-10-05','2026-10-09'), 'OCT S2':('2026-10-12','2026-10-16'),
    'OCT S3':('2026-10-19','2026-10-23'), 'OCT S4':('2026-10-26','2026-10-30'),
    'NOV S1':('2026-11-02','2026-11-06'), 'NOV S2':('2026-11-09','2026-11-13'),
    'NOV S3':('2026-11-16','2026-11-20'), 'NOV S4':('2026-11-23','2026-11-27'),
    'DIC S1':('2026-11-30','2026-12-04'), 'DIC S2':('2026-12-07','2026-12-11'),
    'DIC S3':('2026-12-14','2026-12-18'), 'DIC S4':('2026-12-21','2026-12-25'),
}


def _semana_actual():
    """Devuelve la etiqueta de la semana del calendario que contiene hoy."""
    hoy = date.today()
    for sem, (ini, fin) in SEMANA_FECHAS.items():
        if date.fromisoformat(ini) <= hoy <= date.fromisoformat(fin):
            return sem
    # Si está fuera del rango, buscar la más próxima
    for sem in SEMANAS_ORDENADAS:
        ini, _ = SEMANA_FECHAS[sem]
        if date.fromisoformat(ini) >= hoy:
            return sem
    return SEMANAS_ORDENADAS[-1]


@login_required
def semana_view(request):
    semana_actual = _semana_actual()
    semana_param  = request.GET.get('semana', '').strip()
    colab_param   = request.GET.get('colaborador', '').strip()
    semana_inicial = semana_param if semana_param in SEMANAS_ORDENADAS else semana_actual
    procedimientos = list(
        Procedimiento.objects.values_list('nombre', flat=True).order_by('nombre')
    )
    return render(request, 'dashboard/semana.html', {
        'modulo_activo': 'actividades',
        'semanas': SEMANAS_ORDENADAS,
        'semana_actual': semana_actual,
        'semana_inicial': semana_inicial,
        'colaborador_inicial': colab_param,
        'semana_fechas': SEMANA_FECHAS,
        'colaboradores': sorted(set(
            Actividad.objects.values_list('obligacion__colaborador__nombre', flat=True)
        )),
        'procedimientos': procedimientos,
    })


@login_required
def semana_data(request):
    semana = request.GET.get('semana', _semana_actual())
    colaborador = request.GET.get('colaborador', '').strip()
    procedimiento = request.GET.get('procedimiento', '').strip()

    base_qs = Actividad.objects.select_related(
        'obligacion__colaborador__procedimiento'
    ).order_by('obligacion__colaborador__nombre', 'orden')

    if colaborador:
        base_qs = base_qs.filter(obligacion__colaborador__nombre=colaborador)
    if procedimiento:
        base_qs = base_qs.filter(obligacion__colaborador__procedimiento__nombre=procedimiento)
    qs = [a for a in base_qs if semana in (a.semanas_activas or [])]

    # Agrupar por colaborador
    grupos = {}
    for a in qs:
        c = a.obligacion.colaborador.nombre
        if c not in grupos:
            grupos[c] = []
        grupos[c].append({
            'id':           a.pk,
            'actividad_id': a.actividad_id,
            'descripcion':  a.descripcion,
            'obligacion':   a.obligacion.descripcion,
            'estado':       a.estado,
            'progreso':     a.progreso,
            'fecha_inicio': a.fecha_inicio.isoformat(),
            'fecha_fin':    a.fecha_fin.isoformat(),
            'total_semanas': len(a.semanas_activas),
            'semana_num': (a.semanas_activas.index(semana) + 1) if semana in a.semanas_activas else 0,
        })

    fechas = SEMANA_FECHAS.get(semana, ('', ''))

    return JsonResponse({
        'semana':              semana,
        'fecha_inicio':        fechas[0],
        'fecha_fin':           fechas[1],
        'total_colaboradores': len(grupos),
        'total_actividades':   sum(len(v) for v in grupos.values()),
        'grupos':              grupos,
    })


# ============================================================
#  VISTA RESUMEN GENERAL — KPIs semanales para subdirector
# ============================================================

def _calcular_resumen(semana):
    """Calcula KPIs y tabla por colaborador para una semana dada."""
    base_qs = Actividad.objects.select_related(
        'obligacion__colaborador'
    ).order_by('obligacion__colaborador__nombre', 'orden')
    qs = [a for a in base_qs if semana in (a.semanas_activas or [])]

    total       = len(qs)
    completadas = sum(1 for a in qs if a.estado == 'completada')
    en_curso    = sum(1 for a in qs if a.estado == 'en_curso')
    pendientes  = sum(1 for a in qs if a.estado == 'pendiente')
    bloqueadas  = sum(1 for a in qs if a.estado == 'bloqueada')
    avance_general = round(sum(a.progreso for a in qs) / total) if total > 0 else 0

    grupos = {}
    for a in qs:
        c = a.obligacion.colaborador.nombre
        if c not in grupos:
            grupos[c] = {'compromisos': 0, 'completadas': 0, 'en_curso': 0,
                         'pendientes': 0, 'bloqueadas': 0, 'progreso_sum': 0}
        g = grupos[c]
        g['compromisos']  += 1
        g['progreso_sum'] += a.progreso
        if a.estado == 'completada':  g['completadas'] += 1
        elif a.estado == 'en_curso':  g['en_curso']    += 1
        elif a.estado == 'pendiente': g['pendientes']  += 1
        elif a.estado == 'bloqueada': g['bloqueadas']  += 1

    por_colaborador = []
    for nombre, g in sorted(grupos.items()):
        avance = round(g['progreso_sum'] / g['compromisos']) if g['compromisos'] > 0 else 0
        por_colaborador.append({
            'colaborador': nombre,
            'compromisos': g['compromisos'],
            'completadas': g['completadas'],
            'en_curso':    g['en_curso'],
            'pendientes':  g['pendientes'],
            'bloqueadas':  g['bloqueadas'],
            'avance':      avance,
        })

    fechas = SEMANA_FECHAS.get(semana, ('', ''))
    return {
        'semana':            semana,
        'fecha_inicio':      fechas[0],
        'fecha_fin':         fechas[1],
        'total_compromisos': total,
        'avance_general':    avance_general,
        'completadas':       completadas,
        'en_curso':          en_curso,
        'pendientes':        pendientes,
        'bloqueadas':        bloqueadas,
        'por_colaborador':   por_colaborador,
    }


@login_required
def resumen_view(request):
    semana_actual = _semana_actual()
    idx = SEMANAS_ORDENADAS.index(semana_actual) if semana_actual in SEMANAS_ORDENADAS else 0
    semana_default = SEMANAS_ORDENADAS[max(0, idx - 1)]

    # Pre-calcular datos iniciales para evitar depender del fetch
    todas = Actividad.objects.select_related('obligacion__colaborador').all()
    acts_semana = [a for a in todas if semana_default in (a.semanas_activas or [])]
    total = len(acts_semana)
    completadas = sum(1 for a in acts_semana if a.estado == 'completada')
    en_curso = sum(1 for a in acts_semana if a.estado == 'en_curso')
    pendientes = sum(1 for a in acts_semana if a.estado == 'pendiente')
    bloqueadas = sum(1 for a in acts_semana if a.estado == 'bloqueada')
    avance = round(sum(a.progreso for a in acts_semana) / total, 1) if total else 0

    por_colab = {}
    for a in acts_semana:
        nombre = a.obligacion.colaborador.nombre
        if nombre not in por_colab:
            por_colab[nombre] = {'total': 0, 'completadas': 0, 'en_curso': 0,
                                 'pendientes': 0, 'bloqueadas': 0, 'progreso_sum': 0}
        c = por_colab[nombre]
        c['total'] += 1
        c['progreso_sum'] += a.progreso
        c[a.estado] = c.get(a.estado, 0) + 1

    colaboradores_init = sorted([
        {'nombre': n, 'total': c['total'], 'completadas': c['completadas'],
         'en_curso': c['en_curso'], 'pendientes': c['pendientes'],
         'bloqueadas': c['bloqueadas'],
         'avance': round(c['progreso_sum'] / c['total'], 1) if c['total'] else 0}
        for n, c in por_colab.items()
    ], key=lambda x: x['avance'], reverse=True)

    datos_iniciales = {
        'total': total, 'completadas': completadas, 'en_curso': en_curso,
        'pendientes': pendientes, 'bloqueadas': bloqueadas, 'avance': avance,
        'colaboradores': colaboradores_init,
    }

    procedimientos = list(
        Procedimiento.objects.values_list('nombre', flat=True).order_by('nombre')
    )
    return render(request, 'dashboard/resumen.html', {
        'modulo_activo':  'actividades',
        'semanas':        SEMANAS_ORDENADAS,
        'semana_actual':  semana_actual,
        'semana_default': semana_default,
        'semana_fechas':  SEMANA_FECHAS,
        'datos_iniciales': datos_iniciales,
        'procedimientos': procedimientos,
    })


@login_required
def resumen_data(request):
    semana = request.GET.get('semana', _semana_actual())
    procedimiento = request.GET.get('procedimiento', '').strip()
    todas = Actividad.objects.select_related('obligacion__colaborador__procedimiento').all()
    if procedimiento:
        todas = todas.filter(obligacion__colaborador__procedimiento__nombre=procedimiento)
    acts_semana = [a for a in todas if semana in (a.semanas_activas or [])]

    total = len(acts_semana)
    completadas = sum(1 for a in acts_semana if a.estado == 'completada')
    en_curso = sum(1 for a in acts_semana if a.estado == 'en_curso')
    pendientes = sum(1 for a in acts_semana if a.estado == 'pendiente')
    bloqueadas = sum(1 for a in acts_semana if a.estado == 'bloqueada')
    avance = round(sum(a.progreso for a in acts_semana) / total, 1) if total else 0

    # Datos por colaborador
    por_colab = {}
    for a in acts_semana:
        nombre = a.obligacion.colaborador.nombre
        if nombre not in por_colab:
            por_colab[nombre] = {
                'total': 0, 'completadas': 0, 'en_curso': 0,
                'pendientes': 0, 'bloqueadas': 0, 'progreso_sum': 0,
            }
        c = por_colab[nombre]
        c['total'] += 1
        c['progreso_sum'] += a.progreso
        if a.estado == 'completada':
            c['completadas'] += 1
        elif a.estado == 'en_curso':
            c['en_curso'] += 1
        elif a.estado == 'pendiente':
            c['pendientes'] += 1
        elif a.estado == 'bloqueada':
            c['bloqueadas'] += 1

    colaboradores = []
    for nombre, c in sorted(por_colab.items(), key=lambda x: x[1]['progreso_sum'] / max(x[1]['total'], 1), reverse=True):
        avg = round(c['progreso_sum'] / c['total'], 1) if c['total'] else 0
        colaboradores.append({
            'nombre': nombre,
            'total': c['total'],
            'completadas': c['completadas'],
            'en_curso': c['en_curso'],
            'pendientes': c['pendientes'],
            'bloqueadas': c['bloqueadas'],
            'avance': avg,
        })

    return JsonResponse({
        'total': total,
        'completadas': completadas,
        'en_curso': en_curso,
        'pendientes': pendientes,
        'bloqueadas': bloqueadas,
        'avance': avance,
        'colaboradores': colaboradores,
    })


# ============================================================
#  MI CRONOGRAMA — vista self-service para colaboradores
# ============================================================

@login_required
def mi_cronograma_view(request):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.colaborador:
        return redirect('/actividades/')

    colaborador = perfil.colaborador
    semana_actual = _semana_actual()

    return render(request, 'dashboard/mi_cronograma.html', {
        'modulo_activo': 'mi_cronograma',
        'colaborador': colaborador,
        'semana_actual': semana_actual,
        'semanas': SEMANAS_ORDENADAS,
        'semana_fechas': SEMANA_FECHAS,
    })


@login_required
def mi_cronograma_data(request):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.colaborador:
        return JsonResponse({'error': 'Sin colaborador vinculado'}, status=403)

    colaborador = perfil.colaborador
    semana = request.GET.get('semana', '').strip()

    qs = Actividad.objects.filter(
        obligacion__colaborador=colaborador
    ).select_related('obligacion', 'proyecto').order_by('orden', 'fecha_inicio')

    if semana:
        actividades = [a for a in qs if semana in (a.semanas_activas or [])]
    else:
        actividades = list(qs)

    hoy = date.today()
    tasks = []
    for a in actividades:
        estado_visual = a.estado
        if a.estado == 'pendiente':
            if a.fecha_fin < hoy:
                estado_visual = 'vencida'
            elif a.fecha_inicio <= hoy <= a.fecha_fin:
                estado_visual = 'en_curso'

        tasks.append({
            'id':              a.pk,
            'obligacion':      a.obligacion.descripcion,
            'actividad_id':    a.actividad_id,
            'descripcion':     a.descripcion,
            'fecha_inicio':    a.fecha_inicio.isoformat(),
            'fecha_fin':       a.fecha_fin.isoformat(),
            'progreso':        a.progreso,
            'estado':          a.estado,
            'estado_visual':   estado_visual,
            'evidencia':       a.evidencia or '',
            'semanas_activas': a.semanas_activas or [],
        })

    return JsonResponse({
        'colaborador': colaborador.nombre,
        'total': len(tasks),
        'tasks': tasks,
    })


@login_required
@require_POST
def mi_actividad_update(request, pk):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.colaborador:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        a = Actividad.objects.select_related('obligacion__colaborador').get(pk=pk)
    except Actividad.DoesNotExist:
        return JsonResponse({'error': 'No encontrada'}, status=404)

    if a.obligacion.colaborador_id != perfil.colaborador_id:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = json.loads(request.body)

    if 'progreso' in data:
        a.progreso = max(0, min(100, int(data['progreso'])))
    if 'estado' in data and data['estado'] in ('pendiente', 'en_curso', 'completada', 'bloqueada'):
        # Bloquear cierre sin evidencia
        if data['estado'] == 'completada' and a.evidencias.count() == 0:
            return JsonResponse({
                'ok': False,
                'error': 'No puedes marcar como completada sin adjuntar al menos una evidencia.',
            }, status=400)
        a.estado = data['estado']
    if 'evidencia' in data:
        a.evidencia = str(data['evidencia'])[:2000]

    a.save()
    return JsonResponse({'ok': True})


# ============================================================
#  EVIDENCIAS — upload de archivos por actividad
# ============================================================

@login_required
def evidencias_lista(request, actividad_pk):
    """Lista evidencias de una actividad (GET)."""
    try:
        a = Actividad.objects.get(pk=actividad_pk)
    except Actividad.DoesNotExist:
        return JsonResponse({'error': 'Actividad no encontrada'}, status=404)

    evidencias = a.evidencias.all()
    data = [{
        'id':          e.pk,
        'archivo_url': e.archivo.url if e.archivo else '',
        'nombre':      e.nombre_archivo,
        'comentario':  e.comentario,
        'creado_por':  e.creado_por.username if e.creado_por else '',
        'creado_en':   e.creado_en.strftime('%d/%m/%Y %H:%M') if e.creado_en else '',
    } for e in evidencias]

    return JsonResponse({'evidencias': data, 'total': len(data)})


@login_required
@require_POST
def evidencia_subir(request, actividad_pk):
    """Sube una evidencia (archivo + comentario) a una actividad."""
    perfil = getattr(request.user, 'perfil', None)

    try:
        a = Actividad.objects.select_related('obligacion__colaborador').get(pk=actividad_pk)
    except Actividad.DoesNotExist:
        return JsonResponse({'error': 'Actividad no encontrada'}, status=404)

    # Solo el dueño o un admin pueden subir
    # _es_admin importado a nivel de módulo
    es_dueno = perfil and perfil.colaborador_id == a.obligacion.colaborador_id
    if not es_dueno and not _es_admin(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'error': 'No se recibio archivo'}, status=400)

    # Limitar tamaño (10MB)
    if archivo.size > 10 * 1024 * 1024:
        return JsonResponse({'error': 'El archivo no puede superar 10MB'}, status=400)

    comentario = request.POST.get('comentario', '').strip()

    ev = EvidenciaActividad.objects.create(
        actividad=a,
        archivo=archivo,
        comentario=comentario,
        creado_por=request.user,
    )

    return JsonResponse({
        'ok': True,
        'id': ev.pk,
        'nombre': ev.nombre_archivo,
        'archivo_url': ev.archivo.url,
    })


@login_required
@require_POST
def evidencia_eliminar(request, pk):
    """Elimina una evidencia (solo el creador o admin)."""
    try:
        ev = EvidenciaActividad.objects.select_related('actividad__obligacion__colaborador').get(pk=pk)
    except EvidenciaActividad.DoesNotExist:
        return JsonResponse({'error': 'No encontrada'}, status=404)

    # _es_admin importado a nivel de módulo
    if ev.creado_por_id != request.user.pk and not _es_admin(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    ev.archivo.delete(save=False)
    ev.delete()
    return JsonResponse({'ok': True})


# ============================================================
#  REPORTES SEMANALES — tipo daily de Neusi
# ============================================================

@login_required
def reporte_semanal_data(request):
    """GET: lista reportes del colaborador logueado. Filtrable por semana."""
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.colaborador:
        return JsonResponse({'error': 'Sin colaborador vinculado'}, status=403)

    semana = request.GET.get('semana', '').strip()
    qs = ReporteSemanal.objects.filter(colaborador=perfil.colaborador)
    if semana:
        qs = qs.filter(semana=semana)

    reportes = [{
        'id':            r.pk,
        'semana':        r.semana,
        'que_hizo':      r.que_hizo,
        'impedimentos':  r.impedimentos,
        'creado_en':     r.creado_en.strftime('%d/%m/%Y %H:%M') if r.creado_en else '',
        'actualizado_en': r.actualizado_en.strftime('%d/%m/%Y %H:%M') if r.actualizado_en else '',
    } for r in qs]

    return JsonResponse({'reportes': reportes})


@login_required
@require_POST
def reporte_semanal_guardar(request):
    """Crea o actualiza el reporte de la semana para el colaborador logueado."""
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.colaborador:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = json.loads(request.body)
    semana = data.get('semana', '').strip()
    que_hizo = data.get('que_hizo', '').strip()

    if not semana or not que_hizo:
        return JsonResponse({'error': 'Semana y reporte son obligatorios'}, status=400)

    reporte, created = ReporteSemanal.objects.update_or_create(
        colaborador=perfil.colaborador,
        semana=semana,
        defaults={
            'que_hizo': que_hizo,
            'impedimentos': data.get('impedimentos', '').strip(),
        },
    )

    return JsonResponse({
        'ok': True,
        'id': reporte.pk,
        'created': created,
    })


@admin_required
def reportes_admin(request):
    """Vista admin: lista todos los reportes semanales, filtrable por semana."""
    semana = request.GET.get('semana', '').strip()
    qs = ReporteSemanal.objects.select_related('colaborador').all()
    if semana:
        qs = qs.filter(semana=semana)

    reportes = [{
        'id':            r.pk,
        'colaborador':   r.colaborador.nombre,
        'semana':        r.semana,
        'que_hizo':      r.que_hizo,
        'impedimentos':  r.impedimentos,
        'creado_en':     r.creado_en.strftime('%d/%m/%Y %H:%M') if r.creado_en else '',
    } for r in qs]

    return JsonResponse({'reportes': reportes, 'total': len(reportes)})
